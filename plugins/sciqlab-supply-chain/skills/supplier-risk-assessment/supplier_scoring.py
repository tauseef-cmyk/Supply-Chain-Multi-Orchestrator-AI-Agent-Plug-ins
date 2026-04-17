"""
Supplier Risk Assessment — Scoring & Excel Report Generator
Reads research JSON, uses Claude API to score each supplier across 6 risk dimensions,
and generates a color-coded Excel report.
"""

import json
import os
import sys
from pathlib import Path

try:
    import anthropic
except ImportError:
    print("ERROR: anthropic not installed. Run: pip install anthropic")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
JSON_INPUT = Path(__file__).parent / "supplier_research_results.json"
EXCEL_OUTPUT = Path(__file__).parent / "supplier_risk_report.xlsx"

RISK_DIMENSIONS = [
    "financial_health",
    "geopolitical_risk",
    "esg_sustainability",
    "supply_continuity",
    "compliance_sanctions",
    "reputation",
]

DIMENSION_LABELS = {
    "financial_health": "Financial Health",
    "geopolitical_risk": "Geopolitical Risk",
    "esg_sustainability": "ESG / Sustainability",
    "supply_continuity": "Supply Continuity",
    "compliance_sanctions": "Compliance / Sanctions",
    "reputation": "Reputation",
}

DIMENSION_WEIGHTS = {
    "financial_health": 0.25,
    "geopolitical_risk": 0.20,
    "esg_sustainability": 0.15,
    "supply_continuity": 0.20,
    "compliance_sanctions": 0.10,
    "reputation": 0.10,
}

SCORING_PROMPT = """You are a supply chain risk analyst. Analyze the following market research about a supplier and score them on each risk dimension.

**Supplier:** {supplier_name}
**Country:** {country}
**Category:** {category}
**Website:** {website}

**Research Data:**
{research_text}

Score each of the following dimensions on a scale of 1-10 (1 = very low risk, 10 = very high risk).
Provide a brief evidence summary (2-3 sentences) for each score.

Respond in this exact JSON format only — no markdown, no extra text:
{{
  "financial_health": {{"score": <int>, "evidence": "<text>"}},
  "geopolitical_risk": {{"score": <int>, "evidence": "<text>"}},
  "esg_sustainability": {{"score": <int>, "evidence": "<text>"}},
  "supply_continuity": {{"score": <int>, "evidence": "<text>"}},
  "compliance_sanctions": {{"score": <int>, "evidence": "<text>"}},
  "reputation": {{"score": <int>, "evidence": "<text>"}},
  "overall_recommendation": "<1-2 sentence recommendation>"
}}"""


def format_research_text(research: dict) -> str:
    """Format research data into readable text for the LLM prompt."""
    parts = []
    for dim_key in RISK_DIMENSIONS:
        dim_data = research.get(dim_key, {})
        label = dim_data.get("label", dim_key)
        answer = dim_data.get("answer", "No data available.")
        sources = dim_data.get("sources", [])
        source_titles = [s.get("title", "Unknown") for s in sources[:3]]

        parts.append(f"### {label}")
        parts.append(f"Summary: {answer}")
        if source_titles:
            parts.append(f"Sources: {', '.join(source_titles)}")
        parts.append("")

    return "\n".join(parts)


def score_supplier(client: anthropic.Anthropic, supplier_data: dict) -> dict:
    """Use Claude to analyze research and produce risk scores."""
    name = supplier_data["supplier_name"]
    print(f"  Scoring: {name}...")

    research_text = format_research_text(supplier_data.get("research", {}))

    prompt = SCORING_PROMPT.format(
        supplier_name=name,
        country=supplier_data["country"],
        category=supplier_data["category"],
        website=supplier_data["website"],
        research_text=research_text,
    )

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )

        response_text = message.content[0].text.strip()
        # Handle potential markdown code block wrapping
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1]
            response_text = response_text.rsplit("```", 1)[0].strip()

        scores = json.loads(response_text)
        print(f"    -> Done (Overall weighted: {calculate_overall_score(scores):.1f})")
        return scores

    except Exception as e:
        print(f"    -> ERROR scoring {name}: {e}")
        return {
            dim: {"score": 5, "evidence": f"Scoring failed: {e}"}
            for dim in RISK_DIMENSIONS
        } | {"overall_recommendation": "Manual review required — automated scoring failed."}


def calculate_overall_score(scores: dict) -> float:
    """Calculate weighted average risk score."""
    total = 0.0
    for dim, weight in DIMENSION_WEIGHTS.items():
        dim_score = scores.get(dim, {}).get("score", 5)
        total += dim_score * weight
    return round(total, 1)


def get_risk_rating(score: float) -> str:
    if score <= 3.0:
        return "Low"
    elif score <= 6.0:
        return "Medium"
    else:
        return "High"


def get_risk_fill(score: int | float) -> PatternFill:
    """Return color fill based on risk score."""
    if score <= 3:
        return PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # green
    elif score <= 6:
        return PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # amber
    else:
        return PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")  # red


def get_rating_fill(rating: str) -> PatternFill:
    fills = {
        "Low": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "Medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "High": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    }
    return fills.get(rating, PatternFill())


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_summary_sheet(wb: Workbook, scored_suppliers: list[dict]):
    """Create the summary dashboard sheet."""
    ws = wb.active
    ws.title = "Risk Dashboard"

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Supplier Risk Assessment Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated by AI Agent — {len(scored_suppliers)} suppliers analyzed"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers (row 4)
    headers = [
        "Supplier", "Country", "Category",
        "Financial\nHealth", "Geopolitical\nRisk", "ESG /\nSustainability",
        "Supply\nContinuity", "Compliance /\nSanctions", "Reputation",
        "Overall\nScore", "Risk\nRating",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 35

    # Data rows
    for row_idx, supplier in enumerate(scored_suppliers, 5):
        scores = supplier["scores"]
        overall = supplier["overall_score"]
        rating = supplier["risk_rating"]

        data = [
            supplier["supplier_name"],
            supplier["country"],
            supplier["category"],
        ]
        # Add dimension scores
        for dim in RISK_DIMENSIONS:
            data.append(scores.get(dim, {}).get("score", "N/A"))
        data.append(overall)
        data.append(rating)

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

            # Color-code score columns (4-9) and overall (10)
            if 4 <= col <= 10 and isinstance(value, (int, float)):
                cell.fill = get_risk_fill(value)
                cell.font = Font(name="Calibri", size=10, bold=True)

            # Color-code rating column (11)
            if col == 11:
                cell.fill = get_rating_fill(str(value))
                cell.font = Font(name="Calibri", size=10, bold=True)

    # Column widths
    widths = [28, 16, 24, 12, 12, 14, 12, 14, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = "D5"


def build_detail_sheet(wb: Workbook, supplier: dict):
    """Create a detail sheet for one supplier."""
    name = supplier["supplier_name"]
    # Sheet name max 31 chars
    sheet_name = name[:28] + "..." if len(name) > 31 else name
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Risk Assessment: {name}"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Supplier info
    info = [
        ("Country:", supplier["country"]),
        ("Category:", supplier["category"]),
        ("Website:", supplier["website"]),
        ("Overall Score:", supplier["overall_score"]),
        ("Risk Rating:", supplier["risk_rating"]),
    ]
    for row_idx, (label, value) in enumerate(info, 3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.font = BODY_FONT
        if label == "Risk Rating:":
            cell.fill = get_rating_fill(str(value))
            cell.font = Font(name="Calibri", bold=True, size=10)
        if label == "Overall Score:":
            cell.fill = get_risk_fill(value)
            cell.font = Font(name="Calibri", bold=True, size=10)

    # Dimension detail table
    row = 9
    detail_headers = ["Risk Dimension", "Score (1-10)", "Evidence"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    scores = supplier["scores"]
    for dim in RISK_DIMENSIONS:
        row += 1
        dim_data = scores.get(dim, {})
        score_val = dim_data.get("score", "N/A")
        evidence = dim_data.get("evidence", "No data")

        ws.cell(row=row, column=1, value=DIMENSION_LABELS[dim]).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER

        score_cell = ws.cell(row=row, column=2, value=score_val)
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.border = THIN_BORDER
        if isinstance(score_val, (int, float)):
            score_cell.fill = get_risk_fill(score_val)
            score_cell.font = Font(name="Calibri", bold=True, size=10)

        evidence_cell = ws.cell(row=row, column=3, value=evidence)
        evidence_cell.font = BODY_FONT
        evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")
        evidence_cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 45

    # Recommendation
    row += 2
    ws.cell(row=row, column=1, value="Recommendation:").font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=3)
    rec_cell = ws.cell(row=row + 1, column=1, value=scores.get("overall_recommendation", "N/A"))
    rec_cell.font = Font(name="Calibri", size=10, italic=True)
    rec_cell.alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 65


def generate_excel_report(scored_suppliers: list[dict]):
    """Generate the full Excel report."""
    wb = Workbook()
    build_summary_sheet(wb, scored_suppliers)

    for supplier in scored_suppliers:
        build_detail_sheet(wb, supplier)

    wb.save(EXCEL_OUTPUT)
    print(f"\n  Excel report saved: {EXCEL_OUTPUT.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY environment variable.")
        sys.exit(1)

    if not JSON_INPUT.exists():
        print(f"ERROR: Research file not found: {JSON_INPUT}")
        print("  Run supplier_analysis.py first.")
        sys.exit(1)

    with open(JSON_INPUT, encoding="utf-8") as f:
        research_data = json.load(f)

    print(f"Loaded research for {len(research_data)} suppliers")
    client = anthropic.Anthropic(api_key=api_key)

    scored_suppliers = []
    for i, supplier_data in enumerate(research_data, 1):
        print(f"\n[{i}/{len(research_data)}]", end="")
        scores = score_supplier(client, supplier_data)
        overall = calculate_overall_score(scores)
        rating = get_risk_rating(overall)

        scored_suppliers.append(
            {
                "supplier_name": supplier_data["supplier_name"],
                "website": supplier_data["website"],
                "country": supplier_data["country"],
                "category": supplier_data["category"],
                "scores": scores,
                "overall_score": overall,
                "risk_rating": rating,
            }
        )

    # Sort by overall score descending (highest risk first)
    scored_suppliers.sort(key=lambda x: x["overall_score"], reverse=True)

    # Print console summary
    print(f"\n{'='*80}")
    print(f"  SUPPLIER RISK ASSESSMENT SUMMARY")
    print(f"{'='*80}")
    print(f"  {'Supplier':<30} {'Country':<15} {'Score':>6}  {'Rating':<8}")
    print(f"  {'-'*30} {'-'*15} {'-'*6}  {'-'*8}")
    for s in scored_suppliers:
        print(f"  {s['supplier_name']:<30} {s['country']:<15} {s['overall_score']:>6.1f}  {s['risk_rating']:<8}")
    print(f"{'='*80}")

    generate_excel_report(scored_suppliers)
    print(f"\n  Done! Open {EXCEL_OUTPUT.name} to review the full report.")


if __name__ == "__main__":
    main()
